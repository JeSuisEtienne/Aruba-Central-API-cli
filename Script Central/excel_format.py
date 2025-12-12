"""
Module pour formater les fichiers Excel générés.

Ce module applique un formatage professionnel aux feuilles Excel :
- Bordures sur toutes les cellules
- En-têtes en gras
- Ajustement automatique de la largeur des colonnes
- Filtres automatiques sur la première ligne
"""

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side


def formater_excel(writer, nom_feuille):
    """
    Applique un formatage professionnel à une feuille Excel.
    
    Cette fonction améliore la lisibilité du fichier Excel en ajoutant :
    - Des bordures sur toutes les cellules
    - Des en-têtes en gras
    - Un ajustement automatique de la largeur des colonnes
    - Des filtres automatiques sur la première ligne
    
    Parameters
    ----------
    writer : pd.ExcelWriter
        Objet ExcelWriter de pandas utilisé pour écrire dans le fichier Excel
    nom_feuille : str
        Nom de la feuille à formater (doit correspondre à une feuille existante)
    """
    # Récupération de la feuille de travail (worksheet) correspondant au nom fourni
    ws = writer.sheets[nom_feuille]

    # Définition d'un style de bordure fine pour toutes les cellules
    # Cela améliore la lisibilité en séparant visuellement chaque cellule
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Parcours de toutes les cellules de la feuille
    for row in ws.iter_rows():
        for cell in row:
            # Application des bordures à chaque cellule
            cell.border = thin_border
            # Mise en gras de la première ligne (en-têtes)
            if cell.row == 1:
                cell.font = Font(bold=True)

    # Ajustement automatique de la largeur de chaque colonne
    # La largeur est calculée en fonction du contenu le plus long de la colonne
    for col in ws.columns:
        max_length = 0
        # Conversion du numéro de colonne en lettre (ex: 1 -> A, 2 -> B)
        col_letter = get_column_letter(col[0].column)
        
        # Parcours de toutes les cellules de la colonne pour trouver la longueur maximale
        for cell in col:
            try:
                if cell.value:
                    # Calcul de la longueur du contenu (converti en string)
                    max_length = max(max_length, len(str(cell.value)))
            except:
                # Ignorer les erreurs potentielles (cellules vides, formats spéciaux, etc.)
                pass
        
        # Définition de la largeur de la colonne = longueur max + 2 caractères de marge
        ws.column_dimensions[col_letter].width = max_length + 2

    # Ajout des filtres automatiques sur la première ligne
    # Les filtres permettent de trier et filtrer les données directement dans Excel
    # On vérifie qu'il y a au moins 2 lignes (en-tête + données) avant d'ajouter les filtres
    if ws.max_row > 1:
        # Définition de la plage de filtres : de A1 à la dernière cellule de données
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"